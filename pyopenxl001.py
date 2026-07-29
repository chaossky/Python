from openpyxl import Workbook, load_workbook

wb=Workbook()
ws=wb.active
ws.title="매출데이터"

ws['A1']="상품명"
ws['B1']="수량"
ws['C1']="단가"
ws['D1']="합계"

ws.append(["사과",10,500,"=B2*C2"])
ws.append(["바나나",5,300,"=B3*C3"])


wb.save("sales_fruits.xlsx")

wb2=load_workbook("sales_fruits.xlsx")
ws2=wb2["매출데이터"]
print(ws2["D2"].value)
