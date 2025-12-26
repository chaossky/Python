import calendar
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== 설정 =====
YEAR = 2026
FILENAME = f"Calendar_{YEAR}.xlsx"

# 한국 공휴일(예시)
KOREA_HOLIDAYS = [
    (1, 1, "신정"),
    (2, 16, "설연휴"),
    (2, 17, "설날"),
    (2, 18, "설연휴"),
    (3, 1, "3·1절"),
    (3, 2, "대체공휴일"),
    (3, 6, "나의 생일"),
    (5, 5, "어린이날"),
    (5, 25, "대체공휴일"),
    (6, 3, "지방선거"),
    (6, 6, "현충일"),
    (8, 15, "광복절"),
    (8, 17, "대체공휴일"),
    (9, 24, "추석연휴"),
    (9, 25, "추석"),
    (9, 28, "추석연휴"),
    (10, 3, "개천절"),
    (10, 5, "대체공휴일"),
    (10, 9, "한글날"),
    (12, 25, "성탄절"),
]

# ===== 스타일 =====
thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

title_font = Font(size=16, bold=True)
weekday_font = Font(size=11, bold=True)

# ===== 유틸 =====
def is_holiday(month, day):
    for m, d, name in KOREA_HOLIDAYS:
        if m == month and d == day:
            return name
    return None

def set_col_widths(ws):
    widths = [6] + [16] * 7
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def set_row_heights(ws):
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    for r in range(3, 3 + 6):
        ws.row_dimensions[r].height = 80

def make_month_sheet(wb, year, month):
    ws = wb.create_sheet(title=f"{month}월")

    # 제목
    ws.merge_cells("A1:H1")
    ws["A1"].value = f"{year}년 {month}월"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = border

    # 헤더 (주 번호 + 일~토)
    headers = ["Week", "Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = weekday_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    set_col_widths(ws)
    set_row_heights(ws)

    # 달력 생성: 일요일 시작
    cal = calendar.Calendar(firstweekday=6)
    month_matrix = cal.monthdayscalendar(year, month)

    # 각 주를 채우기
    start_row = 3
    for week_index, week in enumerate(month_matrix):
        row = start_row + week_index

        # 주 번호
        first_day_in_week = None
        for i, day in enumerate(week):
            if day != 0:
                first_day_in_week = date(year, month, day)
                break
        iso_week = first_day_in_week.isocalendar()[1] if first_day_in_week else ""
        ws.cell(row=row, column=1, value=iso_week).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).border = border

        # 날짜 셀 채우기
        for i, day in enumerate(week):
            col = 2 + i
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="right", vertical="top")  # 오른쪽 상단 정렬

            if day == 0:
                cell.value = ""
                continue

            holiday_name = is_holiday(month, day)

            # 기본 날짜
            cell.value = f"{day}"
            cell.font = Font(size=20, bold=True)

            # 공휴일 또는 일요일 → 빨강
            if holiday_name or i == 0:
                cell.font = Font(size=20, bold=True, color="FF0000")
                if holiday_name:
                    cell.value = f"{day}\n{holiday_name}"

            # 토요일 → 파랑
            elif i == 6:
                cell.font = Font(size=20, bold=True, color="0000FF")

    # 테두리 마무리
    max_rows = start_row + len(month_matrix) - 1
    for r in range(1, max_rows + 1):
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = border

def build_workbook(year):
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for m in range(1, 13):
        make_month_sheet(wb, year, m)

    return wb

if __name__ == "__main__":
    wb = build_workbook(YEAR)
    wb.save(FILENAME)
    print(f"엑셀 파일 저장 완료: {FILENAME}")
