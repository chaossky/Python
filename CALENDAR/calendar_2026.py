import calendar
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ===== 설정 =====
YEAR = 2026
FILENAME = f"Calendar_{YEAR}.xlsx"

# 한국 공휴일(예시) — 필요시 정확한 날짜로 수정/추가하세요.
# 형식: (월, 일, "이름")
KOREA_HOLIDAYS = [
    (1, 1, "신정"),
    (2, 16, "설연휴"),
    (2, 17, "설날"),
    (2, 18, "설연휴"),
    (3, 1, "3·1절"),
    (3, 2, "대체공휴일"),
    (3, 6,"나의 생일"),
    (5, 5, "어린이날"),
    (5, 25, "대체공휴일"),  # 음력 변동, 확인 필요
    (6,3,"지방선거"),
    (6, 6, "현충일"),
    (8, 15, "광복절"),
    (8, 17, "대체공휴일"),
    (9, 24, "추석연휴"),       # 음력 변동, 확인 필요
    (9, 25, "추석"),
    (9, 28, "추석연휴"),
    (10, 3, "개천절"),
    (10, 5, "대체공휴일"),
    (10, 9, "한글날"),
    (12, 25, "성탄절"),
]
HOLIDAY_COLOR = "FFC7CE"   # 연분홍
WEEKEND_COLOR = "DAEEF3"   # 옅은 청록
HEADER_COLOR = "D9D9D9"    # 회색 헤더

# ===== 스타일 =====
thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

title_font = Font(size=11, bold=True)
weekday_font = Font(size=11, bold=True)
day_font = Font(size=16,bold=True)
holiday_font = Font(color="9C0006", bold=True)  # 진한 빨강

center = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ===== 유틸 =====
def is_holiday(month, day):
    for m, d, name in KOREA_HOLIDAYS:
        if m == month and d == day:
            return name
    return None

def set_col_widths(ws):
    # 첫 열은 주 번호, 나머지 열은 날짜
    widths = [6] + [16] * 7
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def set_row_heights(ws):
    # 제목/헤더/주차 행 높이
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    for r in range(3, 3 + 6):  # 최대 6주
        ws.row_dimensions[r].height = 80

def make_month_sheet(wb, year, month):
    ws = wb.create_sheet(title=f"{month}월")

    # 제목
    ws.merge_cells("A1:H1")
    ws["A1"].value = f"{year}년 {month}월"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = PatternFill("solid", HEADER_COLOR)
    ws["A1"].border = border

    # 헤더 (주 번호 + 월,화,수,목,금,토,일)
    headers = ["Week"] + ["Sun","Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = weekday_font
        cell.alignment = center
        cell.fill = PatternFill("solid", HEADER_COLOR)
        cell.border = border

    set_col_widths(ws)
    set_row_heights(ws)

    # 달력 생성: calendar.monthcalendar는 주별 리스트(0은 공란)를 반환
    cal = calendar.Calendar(firstweekday=6)  # 0=월요일? 파이썬은 0=월요일 아님, 아래에서 직접 처리
    # Python의 calendar에서는 0=월요일이 아니라 0=월요일로 문서화되어 있지만
    # 실제 weekday() 기준 0=월요일, 6=일요일. 여기서는 header에 "월~일"로 둡니다.
    month_matrix = cal.monthdayscalendar(year, month)

    # 각 주를 채우기
    start_row = 3
    for week_index, week in enumerate(month_matrix):
        row = start_row + week_index
        # 주 번호(ISO Week Number) — 월의 월~일 중 월-일 범위의 목요일 기준 ISO 주 계산
        # 간단하게 해당 주의 첫 유효일로 계산
        first_day_in_week = None
        for i, day in enumerate(week):
            if day != 0:
                first_day_in_week = date(year, month, day)
                break
        iso_week = first_day_in_week.isocalendar()[1] if first_day_in_week else ""
        ws.cell(row=row, column=1, value=iso_week).alignment = center
        ws.cell(row=row, column=1).border = border

        # 날짜 셀 채우기 (월=col2 ... 일=col8)
        for i, day in enumerate(week):
            col = 2 + i  # 월=2, 화=3, ..., 일=8
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if day == 0:
                cell.value = ""
                continue

            # 기본 표시: 날짜 숫자
            cell.value = f"{day}"

            # 주말 색 (토: i=5, 일: i=6)
            if i == 0 or i == 6:
                cell.fill = PatternFill("solid", WEEKEND_COLOR)

            # 공휴일 표시
            holiday_name = is_holiday(month, day)
            if holiday_name:
                cell.font = holiday_font
                cell.fill = PatternFill("solid", HOLIDAY_COLOR)
                # 날짜 아래 줄바꿈으로 이름 표기
                cell.value = f"{day}\n{holiday_name}"

    # 격자 테두리 마무리
    max_rows = start_row + len(month_matrix) - 1
    for r in range(1, max_rows + 1):
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = border

def build_workbook(year):
    wb = Workbook()
    # 기본 생성된 첫 시트는 제거
    default = wb.active
    wb.remove(default)

    # 1월~12월 시트 생성
    for m in range(1, 13):
        make_month_sheet(wb, year, m)

    # 요일 이름을 월~일로 유지하기 위해 locale-independent로 구성
    return wb

if __name__ == "__main__":
    wb = build_workbook(YEAR)
    wb.save(FILENAME)
    print(f"엑셀 파일 저장 완료: {FILENAME}")
